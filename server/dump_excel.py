import os
from openpyxl import load_workbook

EXCEL_PATH = r"C:\RF\Dirección General de Educación Tecnológica Industrial y de Servicios\CETIS No. 150 - APASEO EL ALTO\11. SEGUIMIENTO DEL GASTO\P.I. FEB-JUL 26\CE_150 SEGUIMIENTO GASTO FEB-JUL26.xlsx"

def dump_data():
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        sheet = wb.active
        print(f"Sheet Name: {sheet.title}")
        for r in range(40, 60):
            row_vals = [str(sheet.cell(row=r, column=c).value) for c in range(1, 10)]
            print(f"Row {r}: {' | '.join(row_vals)}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    dump_data()
